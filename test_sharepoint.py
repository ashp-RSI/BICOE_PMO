"""
SharePoint Connectivity Test Script (v3)
=========================================
Tries multiple authentication methods including interactive device-code flow.
"""

import os
import sys
from io import BytesIO
from dotenv import load_dotenv

load_dotenv()

SITE_URL = os.getenv("SHAREPOINT_SITE_URL")
USERNAME = os.getenv("SHAREPOINT_USERNAME")
PASSWORD = os.getenv("SHAREPOINT_PASSWORD")
TENANT_ID = os.getenv("TENANT_ID")
FILE_URL = os.getenv("FILE_RELATIVE_URL")
SHEET_NAME = os.getenv("SHEET_NAME", "Sheet1")
SHAREPOINT_HOST = "rsystemsiltd.sharepoint.com"


def try_msal_ropc(username_override=None):
    """MSAL ROPC flow with explicit tenant ID."""
    import msal

    user = username_override or USERNAME
    authority = f"https://login.microsoftonline.com/{TENANT_ID}"
    client_id = "d3590ed6-52b3-4102-aeff-aad2292ab01c"  # Microsoft Office
    scopes = [f"https://{SHAREPOINT_HOST}/.default"]

    app = msal.PublicClientApplication(client_id, authority=authority)
    result = app.acquire_token_by_username_password(user, PASSWORD, scopes=scopes)

    if "access_token" in result:
        return result["access_token"]
    raise Exception(result.get("error_description", result.get("error", "Unknown error")))


def try_device_code_flow():
    """Interactive device-code flow — works with MFA, federation, any setup."""
    import msal

    authority = f"https://login.microsoftonline.com/{TENANT_ID}"
    client_id = "d3590ed6-52b3-4102-aeff-aad2292ab01c"
    scopes = [f"https://{SHAREPOINT_HOST}/.default"]

    app = msal.PublicClientApplication(client_id, authority=authority)
    flow = app.initiate_device_flow(scopes=scopes)

    if "user_code" not in flow:
        raise Exception(f"Device flow initiation failed: {flow.get('error_description', flow)}")

    print(f"\n  *** ACTION REQUIRED ***")
    print(f"  1. Open your browser and go to: {flow['verification_uri']}")
    print(f"  2. Enter this code: {flow['user_code']}")
    print(f"  3. Sign in with your Microsoft account")
    print(f"  Waiting for you to complete sign-in...\n")

    result = app.acquire_token_by_device_flow(flow)

    if "access_token" in result:
        return result["access_token"]
    raise Exception(result.get("error_description", result.get("error", "Unknown error")))


def download_with_token(access_token):
    """Download file using SharePoint REST API with bearer token."""
    import requests

    api_url = f"{SITE_URL}/_api/web/GetFileByServerRelativeUrl('{FILE_URL}')/$value"
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/octet-stream",
    }
    resp = requests.get(api_url, headers=headers)
    if resp.status_code != 200:
        raise Exception(f"HTTP {resp.status_code}: {resp.text[:300]}")
    return BytesIO(resp.content)


def inspect_excel(buffer):
    """Analyze the downloaded Excel file."""
    import openpyxl
    import pandas as pd

    print(f"\n[3/4] Inspecting Excel structure...")
    wb = openpyxl.load_workbook(buffer, read_only=True)
    print(f"  [OK] Sheet names: {wb.sheetnames}")

    for name in wb.sheetnames:
        ws = wb[name]
        row_count = ws.max_row
        col_count = ws.max_column
        first_row = list(ws.iter_rows(min_row=1, max_row=1))
        headers = [cell.value for cell in first_row[0]] if first_row else []
        print(f"\n  Sheet: '{name}'")
        print(f"    Rows: {row_count}, Columns: {col_count}")
        print(f"    Headers: {headers}")
    wb.close()

    print(f"\n[4/4] Loading target sheet '{SHEET_NAME}' with pandas...")
    buffer.seek(0)
    try:
        df = pd.read_excel(buffer, sheet_name=SHEET_NAME, engine="openpyxl")
    except ValueError as e:
        print(f"  [WARN] Sheet '{SHEET_NAME}' not found. Trying first sheet...")
        buffer.seek(0)
        df = pd.read_excel(buffer, sheet_name=0, engine="openpyxl")

    print(f"  [OK] Loaded {len(df)} rows x {len(df.columns)} columns\n")

    print("  Column Details:")
    print("  " + "-" * 56)
    print(f"  {'Column Name':<30} {'Dtype':<15} {'Non-Null'}")
    print("  " + "-" * 56)
    for col in df.columns:
        non_null = df[col].notna().sum()
        print(f"  {str(col):<30} {str(df[col].dtype):<15} {non_null}/{len(df)}")

    print(f"\n  Unique values in key filter columns:")
    found_filter = False
    for col_name in df.columns:
        col_lower = str(col_name).lower().replace("_", " ")
        if any(kw in col_lower for kw in ["sub practice", "billable", "sub_practice", "practice"]):
            found_filter = True
            unique_vals = df[col_name].dropna().unique().tolist()
            print(f"\n  '{col_name}' ({len(unique_vals)} unique):")
            for v in sorted(unique_vals, key=str):
                count = (df[col_name] == v).sum()
                print(f"    - {v} ({count} rows)")

    if not found_filter:
        print("  (No columns matching 'sub practice' or 'billable' found)")
        print("  All columns:", list(df.columns))

    print(f"\n  First 5 rows preview:")
    print(df.head().to_string(index=True))


def main():
    print("=" * 60)
    print("  SharePoint Connectivity Test (v3)")
    print("=" * 60)

    missing = []
    for var in ["SHAREPOINT_SITE_URL", "SHAREPOINT_USERNAME", "SHAREPOINT_PASSWORD", "TENANT_ID", "FILE_RELATIVE_URL"]:
        val = os.getenv(var)
        if not val or val.startswith("your"):
            missing.append(var)
    if missing:
        print(f"\n[ERROR] Set these in .env: {', '.join(missing)}")
        sys.exit(1)

    print(f"\n  Site:   {SITE_URL}")
    print(f"  User:   {USERNAME}")
    print(f"  Tenant: {TENANT_ID}")
    print(f"  File:   {FILE_URL}")

    onmicrosoft_user = USERNAME.split("@")[0] + "@rsystemsiltd.onmicrosoft.com"

    methods = [
        (f"MSAL ROPC with {USERNAME}", lambda: try_msal_ropc()),
        (f"MSAL ROPC with {onmicrosoft_user}", lambda: try_msal_ropc(onmicrosoft_user)),
        ("Device Code Flow (browser sign-in)", try_device_code_flow),
    ]

    access_token = None

    for name, method in methods:
        print(f"\n[1/4] Trying {name}...")
        try:
            access_token = method()
            print(f"  [OK] Authentication successful!")
            break
        except Exception as e:
            err_msg = str(e)
            if len(err_msg) > 200:
                err_msg = err_msg[:200] + "..."
            print(f"  [SKIP] {err_msg}")

    if access_token is None:
        print(f"\n[FAIL] All methods failed.")
        sys.exit(1)

    print(f"\n[2/4] Downloading file...")
    try:
        buffer = download_with_token(access_token)
        file_size_kb = len(buffer.getvalue()) / 1024
        print(f"  [OK] Downloaded ({file_size_kb:.1f} KB)")
    except Exception as e:
        print(f"  [FAIL] Download failed: {e}")
        sys.exit(1)

    try:
        inspect_excel(buffer)
    except Exception as e:
        print(f"  [FAIL] Excel inspection failed: {e}")
        sys.exit(1)

    print("\n" + "=" * 60)
    print("  ALL TESTS PASSED!")
    print("  Auth method that worked will be used in the main app.")
    print("=" * 60)


if __name__ == "__main__":
    main()
