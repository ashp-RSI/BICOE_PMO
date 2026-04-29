import os
import time
import atexit
import logging
from io import BytesIO

import msal
import pandas as pd
import requests
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class SharePointService:
    """Handles SharePoint authentication (device-code flow with token cache),
    Excel download, and write-back via the SharePoint REST API."""

    def __init__(self, config):
        self.site_url = config["SHAREPOINT_SITE_URL"]
        self.tenant_id = config["TENANT_ID"]
        self.file_url = config["FILE_RELATIVE_URL"]
        self.sheet_name = config["SHEET_NAME"]
        self.host = config["SHAREPOINT_HOST"]
        self.client_id = config["MSAL_CLIENT_ID"]
        self.cache_file = config["TOKEN_CACHE_FILE"]
        self._username = config.get("SHAREPOINT_USERNAME")
        self._password = config.get("SHAREPOINT_PASSWORD")
        self._is_azure = config.get("IS_AZURE", False)

        self._token_cache = msal.SerializableTokenCache()
        if os.path.exists(self.cache_file):
            with open(self.cache_file, "r") as f:
                self._token_cache.deserialize(f.read())

        atexit.register(self._save_cache)

        self._app = msal.PublicClientApplication(
            self.client_id,
            authority=f"https://login.microsoftonline.com/{self.tenant_id}",
            token_cache=self._token_cache,
        )
        self._scopes = [f"https://{self.host}/.default"]

    def _save_cache(self):
        if self._token_cache.has_state_changed:
            with open(self.cache_file, "w") as f:
                f.write(self._token_cache.serialize())

    def get_access_token(self):
        """Return a valid access token, using cached accounts first,
        then username/password (ROPC) flow, then device code (local only)."""
        accounts = self._app.get_accounts()
        if accounts:
            result = self._app.acquire_token_silent(self._scopes, account=accounts[0])
            if result and "access_token" in result:
                logger.info("Token acquired from cache (silent)")
                return result["access_token"]

        if self._username and self._password:
            logger.info("No cached token — trying username/password auth for %s",
                        self._username)
            result = self._app.acquire_token_by_username_password(
                self._username, self._password, scopes=self._scopes
            )
            if result and "access_token" in result:
                self._save_cache()
                return result["access_token"]
            logger.warning("Username/password auth failed: %s",
                           result.get("error_description", result.get("error", "Unknown")))

        if self._is_azure:
            raise RuntimeError(
                "No valid token cache found on Azure. "
                "Please upload .token_cache.bin to /home/ via Kudu. "
                "Generate it by running the app locally first."
            )

        flow = self._app.initiate_device_flow(scopes=self._scopes)
        if "user_code" not in flow:
            raise RuntimeError(f"Device flow failed: {flow}")

        print("\n" + "=" * 50)
        print("  SHAREPOINT SIGN-IN REQUIRED")
        print("=" * 50)
        print(f"  1. Open: {flow['verification_uri']}")
        print(f"  2. Enter code: {flow['user_code']}")
        print(f"  3. Sign in with your Microsoft account")
        print("=" * 50 + "\n")

        result = self._app.acquire_token_by_device_flow(flow)
        if "access_token" not in result:
            raise RuntimeError(
                result.get("error_description", result.get("error", "Auth failed"))
            )

        self._save_cache()
        return result["access_token"]

    def _headers(self):
        token = self.get_access_token()
        return {"Authorization": f"Bearer {token}", "Accept": "application/json"}

    def download_excel(self) -> BytesIO:
        url = (
            f"{self.site_url}/_api/web"
            f"/GetFileByServerRelativeUrl('{self.file_url}')/$value"
        )
        headers = self._headers()
        headers["Accept"] = "application/octet-stream"
        resp = requests.get(url, headers=headers, timeout=30)
        resp.raise_for_status()
        buf = BytesIO(resp.content)
        buf.seek(0)
        return buf

    def get_dataframe(self) -> pd.DataFrame:
        buf = self.download_excel()
        df = pd.read_excel(buf, sheet_name=self.sheet_name, engine="openpyxl")
        df.index.name = "row_index"
        return df

    def get_demand_dataframe(self, demand_sheet_name: str) -> pd.DataFrame:
        buf = self.download_excel()
        try:
            df = pd.read_excel(buf, sheet_name=demand_sheet_name, engine="openpyxl")
        except ValueError:
            df = pd.DataFrame()
        df.index.name = "row_index"
        return df

    def _discard_checkout(self) -> None:
        """Undo any existing checkout on the file so it can be overwritten."""
        url = (
            f"{self.site_url}/_api/web"
            f"/GetFileByServerRelativeUrl('{self.file_url}')"
            f"/UndoCheckOut()"
        )
        headers = self._headers()
        headers["Accept"] = "application/json;odata=verbose"
        headers["X-RequestDigest"] = self._get_request_digest()
        try:
            resp = requests.post(url, headers=headers, timeout=15)
            if resp.ok:
                logger.info("Discarded existing checkout on '%s'", self.file_url)
        except Exception:
            pass

    def _upload_workbook(self, output: BytesIO) -> None:
        """Upload a workbook BytesIO back to SharePoint, retrying on 423 Locked."""
        folder_url = self.file_url.rsplit("/", 1)[0]
        filename = self.file_url.rsplit("/", 1)[1]
        upload_url = (
            f"{self.site_url}/_api/web"
            f"/GetFolderByServerRelativeUrl('{folder_url}')"
            f"/Files/add(url='{filename}',overwrite=true)"
        )

        max_retries = 3
        for attempt in range(max_retries):
            headers = self._headers()
            headers["Accept"] = "application/json;odata=verbose"
            headers["Content-Type"] = "application/octet-stream"
            headers["X-RequestDigest"] = self._get_request_digest()

            resp = requests.post(
                upload_url, headers=headers, data=output.getvalue(), timeout=60
            )
            if resp.status_code == 423 and attempt < max_retries - 1:
                logger.warning("File locked (attempt %d/%d), discarding checkout and retrying...",
                               attempt + 1, max_retries)
                self._discard_checkout()
                time.sleep(3)
                continue
            resp.raise_for_status()
            return

    def _get_header_map(self, ws):
        header_map = {}
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value is not None:
                header_map[str(cell.value)] = cell.column
        return header_map

    def update_row(self, row_index: int, updates: dict) -> None:
        """Download Excel, update specific cells in a row, re-upload."""
        buf = self.download_excel()
        wb = load_workbook(buf)
        ws = wb[self.sheet_name]

        header_map = self._get_header_map(ws)
        next_col = ws.max_column + 1
        excel_row = row_index + 2

        for col_name, new_value in updates.items():
            if col_name not in header_map:
                ws.cell(row=1, column=next_col, value=col_name)
                header_map[col_name] = next_col
                next_col += 1
                logger.info("Added missing column '%s' to sheet '%s'",
                            col_name, self.sheet_name)
            col_idx = header_map[col_name]
            ws.cell(row=excel_row, column=col_idx, value=new_value)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        self._upload_workbook(output)
        logger.info("Row %d updated and uploaded to SharePoint", row_index)

    def update_demand_row(self, demand_sheet_name: str, row_index: int,
                          updates: dict) -> None:
        """Update a row in the demand requisition sheet."""
        buf = self.download_excel()
        wb = load_workbook(buf)
        ws = wb[demand_sheet_name]

        header_map = self._get_header_map(ws)
        excel_row = row_index + 2

        for col_name, new_value in updates.items():
            if col_name not in header_map:
                continue
            ws.cell(row=excel_row, column=header_map[col_name], value=new_value)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        self._upload_workbook(output)
        logger.info("Demand row %d updated in sheet '%s'", row_index,
                     demand_sheet_name)

    def map_employee_to_demand(self, demand_sheet_name: str, demand_row_index: int,
                               demand_updates: dict, headcount_row_index: int,
                               headcount_updates: dict) -> None:
        """Atomically update both the demand sheet and head count sheet in one upload."""
        buf = self.download_excel()
        wb = load_workbook(buf)

        ws_demand = wb[demand_sheet_name]
        demand_hmap = self._get_header_map(ws_demand)
        demand_excel_row = demand_row_index + 2
        for col_name, new_value in demand_updates.items():
            if col_name in demand_hmap:
                ws_demand.cell(row=demand_excel_row,
                               column=demand_hmap[col_name], value=new_value)

        ws_hc = wb[self.sheet_name]
        hc_hmap = self._get_header_map(ws_hc)
        next_col = ws_hc.max_column + 1
        hc_excel_row = headcount_row_index + 2
        for col_name, new_value in headcount_updates.items():
            if col_name not in hc_hmap:
                ws_hc.cell(row=1, column=next_col, value=col_name)
                hc_hmap[col_name] = next_col
                next_col += 1
                logger.info("Added missing column '%s' to sheet '%s'",
                            col_name, self.sheet_name)
            ws_hc.cell(row=hc_excel_row, column=hc_hmap[col_name],
                       value=new_value)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        self._upload_workbook(output)
        logger.info("Mapped employee (HC row %d) to demand (row %d)",
                     headcount_row_index, demand_row_index)

    def delete_demand_row(self, demand_sheet_name: str, row_index: int,
                          revert_emp_row_index: int = None) -> None:
        """Delete a row from the demand sheet. If revert_emp_row_index is given,
        revert the employee's Billable/Non Billable back to 'Non-Billable'."""
        buf = self.download_excel()
        wb = load_workbook(buf)
        ws = wb[demand_sheet_name]

        excel_row = row_index + 2
        ws.delete_rows(excel_row)
        logger.info("Deleted row %d (Excel row %d) from sheet '%s'",
                     row_index, excel_row, demand_sheet_name)

        if revert_emp_row_index is not None:
            ws_hc = wb[self.sheet_name]
            hc_hmap = self._get_header_map(ws_hc)
            hc_excel_row = revert_emp_row_index + 2
            if "Billable/Non Billable" in hc_hmap:
                ws_hc.cell(row=hc_excel_row,
                           column=hc_hmap["Billable/Non Billable"],
                           value="Non-Billable")
                logger.info("Reverted employee (HC row %d) to Non-Billable",
                            revert_emp_row_index)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        self._upload_workbook(output)

    def delete_headcount_row(self, row_index: int) -> None:
        """Delete a row from the headcount sheet."""
        buf = self.download_excel()
        wb = load_workbook(buf)
        ws = wb[self.sheet_name]

        excel_row = row_index + 2
        ws.delete_rows(excel_row)
        logger.info("Deleted headcount row %d (Excel row %d) from sheet '%s'",
                     row_index, excel_row, self.sheet_name)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        self._upload_workbook(output)

    def backfill_billable_columns(self, columns: list) -> int:
        """For every row where Billable/Non Billable == 'Billable', set the
        given columns to 'Yes' if they are currently empty. Returns the count
        of rows updated."""
        buf = self.download_excel()
        wb = load_workbook(buf)
        ws = wb[self.sheet_name]

        header_map = self._get_header_map(ws)

        billable_col = header_map.get("Billable/Non Billable")
        if not billable_col:
            logger.warning("Column 'Billable/Non Billable' not found — skipping backfill")
            return 0

        next_col = ws.max_column + 1
        for col_name in columns:
            if col_name not in header_map:
                ws.cell(row=1, column=next_col, value=col_name)
                header_map[col_name] = next_col
                next_col += 1
                logger.info("Added missing column '%s' during backfill", col_name)

        col_indices = [header_map[c] for c in columns]
        updated = 0

        for row in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=billable_col).value
            if cell_val and str(cell_val).strip() == "Billable":
                row_changed = False
                for ci in col_indices:
                    existing = ws.cell(row=row, column=ci).value
                    if not existing or str(existing).strip() == "":
                        ws.cell(row=row, column=ci, value="Yes")
                        row_changed = True
                if row_changed:
                    updated += 1

        if updated > 0:
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            self._upload_workbook(output)
            logger.info("Backfilled %d Billable rows with %s = Yes", updated, columns)

        return updated

    def clear_column_value(self, column_name: str, value_to_clear: str) -> int:
        """Clear all cells in a column that match a specific value. Returns count of cleared cells."""
        buf = self.download_excel()
        wb = load_workbook(buf)
        ws = wb[self.sheet_name]

        header_map = self._get_header_map(ws)
        col_idx = header_map.get(column_name)
        if not col_idx:
            return 0

        cleared = 0
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value and str(cell.value).strip() == value_to_clear:
                cell.value = None
                cleared += 1

        if cleared > 0:
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            self._upload_workbook(output)
            logger.info("Cleared %d cells with value '%s' from column '%s'",
                        cleared, value_to_clear, column_name)

        return cleared

    def ensure_sheet_exists(self, sheet_name: str, headers: list) -> None:
        """Create a worksheet with headers if it doesn't exist, or add missing
        header columns to an existing sheet."""
        self.ensure_multiple_sheets({sheet_name: headers})

    def ensure_multiple_sheets(self, sheets: dict) -> None:
        """Ensure multiple sheets/columns exist in a single download-upload cycle.
        ``sheets`` maps sheet_name -> list of required headers."""
        buf = self.download_excel()
        wb = load_workbook(buf)
        changed = False

        for sheet_name, headers in sheets.items():
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(title=sheet_name)
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col_idx, value=header)
                changed = True
                logger.info("Created sheet '%s' with headers %s", sheet_name, headers)
            else:
                ws = wb[sheet_name]
                existing = set()
                for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                    if cell.value is not None:
                        existing.add(str(cell.value))
                next_col = ws.max_column + 1
                for header in headers:
                    if header not in existing:
                        ws.cell(row=1, column=next_col, value=header)
                        next_col += 1
                        changed = True
                        logger.info("Added missing column '%s' to sheet '%s'",
                                    header, sheet_name)

        if changed:
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            self._upload_workbook(output)

    def fulfill_external_demand(self, demand_sheet_name: str,
                                demand_row_index: int, demand_updates: dict,
                                headcount_row_data: dict) -> int:
        """Atomically append a new employee to the headcount sheet and update
        the demand row to Fulfilled in a single download-upload cycle.
        Returns the new headcount Excel row number."""
        buf = self.download_excel()
        wb = load_workbook(buf)

        ws_demand = wb[demand_sheet_name]
        demand_hmap = self._get_header_map(ws_demand)
        demand_excel_row = demand_row_index + 2
        for col_name, new_value in demand_updates.items():
            if col_name in demand_hmap:
                ws_demand.cell(row=demand_excel_row,
                               column=demand_hmap[col_name], value=new_value)

        ws_hc = wb[self.sheet_name]
        hc_hmap = self._get_header_map(ws_hc)
        next_col = ws_hc.max_column + 1
        new_row = ws_hc.max_row + 1

        for col_name, value in headcount_row_data.items():
            if col_name not in hc_hmap:
                ws_hc.cell(row=1, column=next_col, value=col_name)
                hc_hmap[col_name] = next_col
                next_col += 1
            ws_hc.cell(row=new_row, column=hc_hmap[col_name], value=value)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        self._upload_workbook(output)
        logger.info("Fulfilled external demand (row %d) and added employee at HC row %d",
                     demand_row_index, new_row)
        return new_row

    def append_row(self, sheet_name: str, row_data: dict, headers: list) -> int:
        """Append a row to the given sheet and return the new Excel row number."""
        buf = self.download_excel()
        wb = load_workbook(buf)
        ws = wb[sheet_name]

        header_map = {}
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value is not None:
                header_map[str(cell.value)] = cell.column

        next_row = ws.max_row + 1

        for col_name, value in row_data.items():
            if col_name in header_map:
                ws.cell(row=next_row, column=header_map[col_name], value=value)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        self._upload_workbook(output)
        logger.info("Appended row %d to sheet '%s'", next_row, sheet_name)
        return next_row

    def append_rows(self, sheet_name: str, rows: list) -> list:
        """Append multiple rows to the given sheet in a single download-upload
        cycle. Unknown columns are created on the header row. Returns the list
        of new Excel row numbers."""
        if not rows:
            return []

        buf = self.download_excel()
        wb = load_workbook(buf)
        ws = wb[sheet_name]

        header_map = self._get_header_map(ws)
        next_col = ws.max_column + 1
        new_rows = []

        for row_data in rows:
            excel_row = ws.max_row + 1
            for col_name, value in row_data.items():
                if col_name not in header_map:
                    ws.cell(row=1, column=next_col, value=col_name)
                    header_map[col_name] = next_col
                    next_col += 1
                    logger.info("Added missing column '%s' to sheet '%s'",
                                col_name, sheet_name)
                ws.cell(row=excel_row, column=header_map[col_name], value=value)
            new_rows.append(excel_row)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        self._upload_workbook(output)
        logger.info("Appended %d rows to sheet '%s'", len(rows), sheet_name)
        return new_rows

    def _get_request_digest(self) -> str:
        url = f"{self.site_url}/_api/contextinfo"
        headers = self._headers()
        headers["Accept"] = "application/json;odata=verbose"
        headers["Content-Length"] = "0"
        resp = requests.post(url, headers=headers, timeout=15)
        resp.raise_for_status()
        return resp.json()["d"]["GetContextWebInformation"]["FormDigestValue"]
